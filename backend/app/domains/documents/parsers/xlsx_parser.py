"""XLSX Excel Spreadsheet Parser for 3-Statements & Financial Data."""

import io
from typing import List
import openpyxl
from app.core.security import sanitize_document_text
from app.domains.documents.parsers.base import ParsedChunk, ParsedDocument


class XlsxParser:
    """Extracts structured financial sheets and table matrices from Excel workbooks."""

    async def parse(self, file_bytes: bytes, filename: str) -> ParsedDocument:
        """Parse raw XLSX workbook bytes into structured tabular chunks."""
        stream = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)

        chunks: List[ParsedChunk] = []
        full_text_parts: List[str] = []
        global_char_offset = 0
        chunk_idx = 0
        table_count = 0

        for sheet_idx, sheetname in enumerate(wb.sheetnames):
            ws = wb[sheetname]
            table_count += 1
            page_num = sheet_idx + 1

            rows_data: List[str] = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                non_empty = [c for c in row if c is not None and str(c).strip() != ""]
                if not non_empty:
                    continue

                formatted_cells = []
                for cell in row:
                    if cell is None:
                        formatted_cells.append("")
                    elif isinstance(cell, float):
                        formatted_cells.append(f"{cell:,.2f}" if abs(cell) >= 1000 else f"{cell:.4f}".rstrip("0").rstrip("."))
                    else:
                        formatted_cells.append(sanitize_document_text(str(cell).strip()))

                rows_data.append(" | ".join(formatted_cells))

            if not rows_data:
                continue

            # Group rows into manageable tabular chunks (~30 rows per chunk)
            batch_size = 30
            for i in range(0, len(rows_data), batch_size):
                batch_rows = rows_data[i : i + batch_size]
                sheet_text = f"### Sheet: {sheetname}\n" + "\n".join(batch_rows)
                full_text_parts.append(sheet_text)

                char_start = global_char_offset
                char_end = char_start + len(sheet_text)
                chunks.append(
                    ParsedChunk(
                        content=sheet_text,
                        page_number=page_num,
                        section_title=f"Spreadsheet: {sheetname}",
                        chunk_index=chunk_idx,
                        token_count=len(sheet_text.split()),
                        char_offset_start=char_start,
                        char_offset_end=char_end,
                        source_type="TABLE",
                        metadata={
                            "sheet_name": sheetname,
                            "row_start": i + 1,
                            "row_end": min(i + batch_size, len(rows_data)),
                            "source_file": filename,
                        },
                    )
                )
                chunk_idx += 1
                global_char_offset = char_end + 1

        wb.close()
        raw_text = "\n\n".join(full_text_parts)

        return ParsedDocument(
            raw_text=raw_text,
            page_count=max(1, len(wb.sheetnames)),
            table_count=table_count,
            chunks=chunks,
            metadata={"filename": filename, "format": "XLSX", "sheets": wb.sheetnames},
        )
